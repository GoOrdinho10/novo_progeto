import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook, load_workbook
import tkinter as tk
from tkinter import messagebox
from tkinter import ttk
import os
from selenium.common.exceptions import NoSuchElementException, TimeoutException

def capturar_dados_climaticos():
    try:
        options = webdriver.ChromeOptions()
        options.add_argument("--disable-blink-features=AutomationControlled")
        options.add_argument("--window-size=1920,1080")
        options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36")
        
        # Teste primeiro sem headless para ver o navegador
        # options.add_argument('--headless=new')
        
        driver = webdriver.Chrome(options=options)
        driver.get("https://www.google.com/search?q=tempo+em+S%C3%A3o+Paulo")

        # Espera explícita para o widget do tempo carregar
        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "div#wob_wc"))
        )

        dados = {
            'temperatura': 'N/A',
            'umidade': 'N/A',
            'vento': 'N/A',
            'chuva': 'N/A'
        }

        try:
            dados['temperatura'] = driver.find_element(By.CSS_SELECTOR, "span#wob_tm").text + "°C"
        except NoSuchElementException:
            print("Elemento da temperatura não encontrado")

        try:
            dados['umidade'] = driver.find_element(By.CSS_SELECTOR, "span#wob_hm").text
        except NoSuchElementException:
            print("Elemento da umidade não encontrado")

        try:
            dados['vento'] = driver.find_element(By.CSS_SELECTOR, "span#wob_ws").text
        except NoSuchElementException:
            print("Elemento do vento não encontrado")

        try:
            dados['chuva'] = driver.find_element(By.CSS_SELECTOR, "span#wob_pp").text
        except NoSuchElementException:
            print("Elemento da chuva não encontrado")

        driver.quit()
        return dados.values()
    
    except Exception as e:
        print(f"Erro geral no navegador: {str(e)}")
        return None, None, None, None

def salvar_dados_planilha(temperatura, umidade, vento, chuva):
    try:
        file_path = "historico_temperatura.xlsx"
        
        if os.path.exists(file_path):
            wb = load_workbook(file_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(["Data/Hora", "Temperatura", "Umidade", "Vento", "Chuva"])
        
        data_hora = time.strftime("%Y-%m-%d %H:%M:%S")
        nova_linha = [data_hora, temperatura, umidade, vento, chuva]
        ws.append(nova_linha)
        wb.save(file_path)
        print(f"Dados salvos: {nova_linha}")
        return True
    
    except PermissionError:
        messagebox.showerror("Erro", "Feche o arquivo Excel antes de salvar")
        return False
    except Exception as e:
        print(f"Erro ao salvar: {str(e)}")
        return False

def buscar_previsao():
    try:
        temperatura, umidade, vento, chuva = capturar_dados_climaticos()
        if None in (temperatura, umidade):
            raise ValueError("Dados incompletos")
            
        if salvar_dados_planilha(temperatura, umidade, vento, chuva):
            messagebox.showinfo("Sucesso", "Dados salvos na planilha!")
        else:
            messagebox.showerror("Erro", "Falha ao gravar dados")
    
    except Exception as e:
        messagebox.showerror("Erro", f"Falha na operação: {str(e)}")

def criar_interface():
    root = tk.Tk()
    root.title("Captador de Clima - São Paulo")
    root.geometry("400x200")
    
    style = ttk.Style()
    style.configure("TButton", padding=10, font=('Arial', 12, 'bold'))
    
    frame = ttk.Frame(root, padding=20)
    frame.pack(expand=True, fill='both')
    
    btn = ttk.Button(frame, text="Capturar Dados", command=buscar_previsao)
    btn.pack(pady=10)
    
    lbl = ttk.Label(frame, text="Verifique o console para detalhes de execução", font=('Arial', 10))
    lbl.pack(pady=5)
    
    root.mainloop()

if __name__ == "__main__":
    criar_interface()